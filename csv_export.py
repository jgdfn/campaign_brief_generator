"""
Takes the model's JSON response (matching schema.py's shape) and writes
the 4-sheet Excel workbook: Campaign Brief, Extraction Log,
Missing Fields, External Suggestions.

Two entry points:
- build_workbook(response_json, output_path): saves to disk
- build_workbook_bytes(response_json): returns bytes, for Streamlit's download_button
"""

import io
import json
import sys
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from schema import BRIEF_FIELDS

HEADER_FILL = PatternFill('solid', start_color='1F2937')
HEADER_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=11)
LABEL_FONT = Font(name='Arial', bold=True, size=10)
VALUE_FONT = Font(name='Arial', size=10)
MISSING_FONT = Font(name='Arial', italic=True, size=10, color='B45309')
WRAP = Alignment(wrap_text=True, vertical='top')
THIN = Side(style='thin', color='D1D5DB')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

SOURCE_COLORS = {
    "Missing": "B45309",
    "Synthesized from MOM": "1D4ED8",
    "Inferred": "7C3AED",
    "Explicit": "15803D",
}


def _style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical='center')
        cell.border = BORDER


def _build_wb_object(brief: dict, suggestions: list) -> Workbook:
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Campaign Brief"
    ws1.cell(row=1, column=1, value="Field")
    ws1.cell(row=1, column=2, value="Response")
    _style_header(ws1, 1, 2)

    r = 2
    for field in BRIEF_FIELDS:
        entry = brief.get(field, {"value": "Missing", "source_type": "Missing"})
        ws1.cell(row=r, column=1, value=field).font = LABEL_FONT
        cell_b = ws1.cell(row=r, column=2, value=entry["value"])
        cell_b.font = MISSING_FONT if entry.get("source_type") == "Missing" else VALUE_FONT
        for c in (1, 2):
            ws1.cell(row=r, column=c).alignment = WRAP
            ws1.cell(row=r, column=c).border = BORDER
        ws1.row_dimensions[r].height = 60
        r += 1

    ws1.column_dimensions['A'].width = 26
    ws1.column_dimensions['B'].width = 80
    ws1.freeze_panes = "A2"

    ws2 = wb.create_sheet("Extraction Log")
    headers = ["Field", "Generated Value", "Source Type", "Evidence"]
    for i, h in enumerate(headers, start=1):
        ws2.cell(row=1, column=i, value=h)
    _style_header(ws2, 1, 4)

    r = 2
    for field in BRIEF_FIELDS:
        entry = brief.get(field, {"value": "Missing", "source_type": "Missing", "evidence": ""})
        ws2.cell(row=r, column=1, value=field).font = LABEL_FONT
        ws2.cell(row=r, column=2, value=entry["value"]).font = VALUE_FONT
        src = entry.get("source_type", "Missing")
        src_cell = ws2.cell(row=r, column=3, value=src)
        src_cell.font = Font(name='Arial', size=10, bold=True,
                              color=SOURCE_COLORS.get(src, "000000"))
        ws2.cell(row=r, column=4, value=entry.get("evidence", "")).font = VALUE_FONT
        for c in range(1, 5):
            ws2.cell(row=r, column=c).alignment = WRAP
            ws2.cell(row=r, column=c).border = BORDER
        ws2.row_dimensions[r].height = 50
        r += 1

    ws2.column_dimensions['A'].width = 22
    ws2.column_dimensions['B'].width = 45
    ws2.column_dimensions['C'].width = 18
    ws2.column_dimensions['D'].width = 50
    ws2.freeze_panes = "A2"

    ws3 = wb.create_sheet("Missing Fields")
    ws3.cell(row=1, column=1, value="Field")
    ws3.cell(row=1, column=2, value="Status")
    _style_header(ws3, 1, 2)

    r = 2
    for field in BRIEF_FIELDS:
        entry = brief.get(field, {})
        if entry.get("source_type") == "Missing":
            ws3.cell(row=r, column=1, value=field).font = LABEL_FONT
            ws3.cell(row=r, column=2, value=entry.get("value", "Missing")).font = MISSING_FONT
            for c in (1, 2):
                ws3.cell(row=r, column=c).border = BORDER
                ws3.cell(row=r, column=c).alignment = WRAP
            r += 1

    ws3.column_dimensions['A'].width = 26
    ws3.column_dimensions['B'].width = 45
    ws3.freeze_panes = "A2"

    ws4 = wb.create_sheet("External Suggestions")
    ws4.cell(row=1, column=1, value="Category")
    ws4.cell(row=1, column=2, value="Suggestion")
    _style_header(ws4, 1, 2)

    r = 2
    for item in suggestions:
        ws4.cell(row=r, column=1, value=item.get("category", "")).font = LABEL_FONT
        ws4.cell(row=r, column=2, value=item.get("suggestion", "")).font = VALUE_FONT
        for c in (1, 2):
            ws4.cell(row=r, column=c).border = BORDER
            ws4.cell(row=r, column=c).alignment = WRAP
        ws4.row_dimensions[r].height = 70
        r += 1

    ws4.column_dimensions['A'].width = 26
    ws4.column_dimensions['B'].width = 80
    ws4.freeze_panes = "A2"

    return wb


def build_workbook(response_json: dict, output_path: str):
    brief = response_json["brief"]
    suggestions = response_json.get("external_suggestions", [])
    wb = _build_wb_object(brief, suggestions)
    wb.save(output_path)


def build_workbook_bytes(response_json: dict) -> bytes:
    """Same as build_workbook but returns bytes in memory, for Streamlit's download_button."""
    brief = response_json["brief"]
    suggestions = response_json.get("external_suggestions", [])
    wb = _build_wb_object(brief, suggestions)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


if __name__ == "__main__":
    with open("../test_outputs/yummiez_simulated_response.json") as f:
        response = json.load(f)
    build_workbook(response, "../test_outputs/Yummiez_Brief_From_Pipeline.xlsx")
    print("Built workbook from simulated model output.")
